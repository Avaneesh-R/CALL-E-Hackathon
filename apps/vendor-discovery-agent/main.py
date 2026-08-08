"""
Vendor Discovery & Outreach - main orchestrator.
Usage:
  python main.py --product "office chairs" --location "Berlin, Germany"
  python main.py --product "hardware supplies" --location "London" --limit 10 --dry-run
"""
import argparse
import json
import sys
from datetime import datetime, timezone
from models import init_db, get_conn, Lead, Campaign, _mask_phone
from discovery import discover_vendors
from script_gen import generate_goal, prompt_client_approval
from caller import (execute_call_pipeline, classify_round1, extract_round2_fields,
                    parse_transcript_to_json, infer_from_transcript)
from scheduler import schedule_followup

ATTRIBUTION = "Data (c) OpenStreetMap contributors (ODbL) - openstreetmap.org/copyright"


def save_lead(conn, vendor: dict, campaign_id: int) -> Lead:
    osm_id = vendor.get("osm_id")
    lead = Lead(
        phone=vendor["phone"],
        campaign_id=campaign_id,
        name=vendor.get("name"),
        category=vendor.get("category"),
        lat=vendor.get("lat"),
        lon=vendor.get("lon"),
        osm_id=osm_id,
        candidate_id=f"osm:{osm_id}" if osm_id else None,
        address=vendor.get("address"),
    )
    lead.save(conn)
    return lead


def update_lead_status(conn, lead_id: int, status: str, call_id: str = None, round_num: int = 1):
    if round_num == 1:
        conn.execute("UPDATE leads SET status=?, round1_call_id=? WHERE id=?",
                     (status, call_id, lead_id))
    else:
        conn.execute("UPDATE leads SET status=?, round2_call_id=? WHERE id=?",
                     (status, call_id, lead_id))
    conn.commit()


def log_call(conn, lead_id: int, call_id: str, round_num: int,
             raw_output: dict, extracted: dict = None):
    conn.execute(
        """INSERT INTO call_logs (lead_id, call_id, round, raw_status_output, extracted_fields)
           VALUES (?,?,?,?,?)""",
        (lead_id, call_id, round_num,
         json.dumps(raw_output), json.dumps(extracted) if extracted else None)
    )
    conn.commit()


def print_results_table(conn, campaign_id: int):
    rows = conn.execute(
        """SELECT l.name, l.phone, l.masked_phone, l.category, l.status,
                  cl.extracted_fields
           FROM leads l
           LEFT JOIN call_logs cl ON cl.lead_id = l.id AND cl.round = 2
           WHERE l.campaign_id = ?
           ORDER BY l.id""",
        (campaign_id,)
    ).fetchall()

    print(f"\n{'='*80}")
    print(f"{'NAME':<25} {'PHONE':<18} {'CATEGORY':<15} {'STATUS':<15} NOTES")
    print(f"{'='*80}")
    for r in rows:
        notes = ""
        if r["extracted_fields"]:
            ef = json.loads(r["extracted_fields"])
            notes = (ef.get("summary") or ef.get("transcript_summary") or str(ef))[:50]
        masked = r["masked_phone"] or _mask_phone(r["phone"])
        print(f"{(r['name'] or '')[:24]:<25} {masked:<18} {(r['category'] or '')[:14]:<15} {r['status']:<15} {notes}")
    print(f"{'='*80}")
    print(f"\n{ATTRIBUTION}")


def export_excel(campaign_id: int, path: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["ID", "Name", "Phone (masked)", "Category", "Status",
               "Interest", "Can Supply", "Price Range", "Timeline",
               "Contact Name", "Next Steps", "Summary", "Candidate ID"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    with get_conn() as conn:
        rows = conn.execute(
            """SELECT l.id, l.name, l.masked_phone, l.phone, l.category, l.status,
                      l.candidate_id,
                      r2log.extracted_fields AS r2_fields,
                      r1log.extracted_fields AS r1_fields
               FROM leads l
               LEFT JOIN call_logs r1log ON r1log.lead_id = l.id AND r1log.round = 1
               LEFT JOIN call_logs r2log ON r2log.lead_id = l.id AND r2log.round = 2
               WHERE l.campaign_id = ?
               ORDER BY l.id""",
            (campaign_id,)
        ).fetchall()

    for row_num, r in enumerate(rows, 2):
        r2 = json.loads(r["r2_fields"]) if r["r2_fields"] else {}
        r1 = json.loads(r["r1_fields"]) if r["r1_fields"] else {}
        summary = r2.get("summary") or r1.get("summary") or ""
        masked = r["masked_phone"] or _mask_phone(r["phone"])
        ws.append([
            r["id"], r["name"] or "", masked, r["category"] or "", r["status"],
            r1.get("interest_level", ""), r2.get("can_supply", ""),
            r2.get("price_range", ""), r2.get("timeline", ""),
            r2.get("contact_name", ""), r2.get("next_steps", ""),
            summary, r["candidate_id"] or "",
        ])

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["L"].width = 40
    wb.save(path)
    print(f"\nExported {len(rows)} leads to {path}")


def run_campaign(product: str, location: str, limit: int,
                 region: str, language: str, dry_run: bool, yes: bool = False,
                 export: str = None, skip_hours_gate: bool = False):
    init_db()

    print(f"\nVendor Discovery & Outreach")
    print(f"Product : {product}")
    print(f"Location: {location}")
    print(f"{'[DRY RUN] ' if dry_run else ''}Discovering vendors via OpenStreetMap...")

    vendors = discover_vendors(product, location, limit)
    if not vendors:
        print("No vendors with phone numbers found. Try a different location or product keyword.")
        sys.exit(1)

    print(f"Found {len(vendors)} vendor(s) with phone numbers.\n")
    for i, v in enumerate(vendors, 1):
        masked = _mask_phone(v["phone"])
        print(f"  {i:>2}. {v['name'][:40]:<40} {masked}  ({v['category']})")

    print(f"\n{ATTRIBUTION}")

    if yes:
        print(f"\nAuto-confirming {len(vendors)} vendors (--yes flag).")
        consent_approved_at = datetime.now(timezone.utc).isoformat()
    else:
        confirm = input(f"\nProceed with these {len(vendors)} vendors? [y/N]: ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            sys.exit(0)
        consent_approved_at = datetime.now(timezone.utc).isoformat()

    with get_conn() as conn:
        campaign = Campaign(
            product_description=product,
            location=location,
            consent_basis="client-reviewed-and-approved",
            consent_approved_at=consent_approved_at,
        )
        campaign.save(conn)
        campaign_id = campaign.id
        conn.commit()

        leads = [save_lead(conn, v, campaign_id) for v in vendors]

    # Generate and get approval for round-1 script
    r1_goal = generate_goal(product, round_num=1)
    r1_goal = r1_goal if yes else prompt_client_approval(r1_goal, round_num=1)

    with get_conn() as conn:
        conn.execute("UPDATE campaigns SET goal_script=? WHERE id=?", (r1_goal, campaign_id))
        conn.commit()

    # -- Round 1 --------------------------------------------------------------
    print(f"\n{'-'*60}")
    print(f"ROUND 1 - Qualification calls ({len(leads)} vendors)")
    print(f"{'-'*60}")

    positives = []
    for lead in leads:
        print(f"\n[{lead.name}]")
        try:
            status_output = execute_call_pipeline(
                lead.phone, r1_goal, region=region, language=language, dry_run=dry_run,
                lat=None if skip_hours_gate else lead.lat,
                lon=None if skip_hours_gate else lead.lon,
            )
            if status_output.get("status") == "SKIPPED":
                with get_conn() as conn:
                    lead.skip_reason = status_output.get("skip_reason")
                    conn.execute("UPDATE leads SET status='skipped', skip_reason=? WHERE id=?",
                                 (lead.skip_reason, lead.id))
                    conn.commit()
                continue
            call_id = status_output.get("run_id") or status_output.get("id") or "dry_run"
            outcome = classify_round1(status_output)
            print(f"  Outcome: {outcome}")

            transcript_lines = parse_transcript_to_json(status_output)
            inference = {}
            if transcript_lines:
                print(f"  Running AI inference on {len(transcript_lines)}-line transcript...")
                try:
                    inference = infer_from_transcript(transcript_lines, product, round_num=1)
                    print(f"  Interest: {inference.get('interest_level','?')} | "
                          f"Rec R2: {inference.get('recommend_round2','?')} | "
                          f"{inference.get('summary','')[:60]}")
                except Exception as ie:
                    print(f"  Inference error: {ie}")

            with get_conn() as conn:
                update_lead_status(conn, lead.id, outcome, call_id, round_num=1)
                log_call(conn, lead.id, call_id, 1, status_output,
                         {"transcript": transcript_lines, **inference} if transcript_lines else inference)

            if outcome == "positive":
                positives.append(lead)
                # Auto-schedule R2 if vendor gave a callback time AND this is an OSM lead
                if lead.osm_id and not dry_run:
                    timeline = inference.get("timeline") or inference.get("next_steps") or ""
                    if timeline and timeline not in ("null", "None", ""):
                        scheduled = schedule_followup(
                            lead_id=lead.id,
                            campaign_id=campaign_id,
                            timeline_text=timeline,
                            product=product,
                            lat=lead.lat,
                            lon=lead.lon,
                        )
                        if not scheduled:
                            print(f"  [Scheduler] Could not parse time from: '{timeline}'")
                    # If no timeline in R1, also check R2 inference from same transcript
                    elif transcript_lines:
                        try:
                            r2_inf = infer_from_transcript(transcript_lines, product, round_num=2)
                            tl2 = r2_inf.get("timeline") or r2_inf.get("next_steps") or ""
                            if tl2 and tl2 not in ("null", "None", ""):
                                schedule_followup(lead_id=lead.id, campaign_id=campaign_id,
                                                  timeline_text=tl2, product=product,
                                                  lat=lead.lat, lon=lead.lon)
                        except Exception:
                            pass
        except Exception as e:
            print(f"  Error: {e}")
            with get_conn() as conn:
                update_lead_status(conn, lead.id, "failed", None, round_num=1)

    if not positives:
        print("\nNo positive responses from round 1.")
        with get_conn() as conn:
            print_results_table(conn, campaign_id)
        if export:
            export_excel(campaign_id, export)
        return

    print(f"\n{len(positives)} positive response(s). Proceeding to round 2.")

    # Generate and get approval for round-2 script
    r2_goal = generate_goal(product, round_num=2)
    r2_goal = r2_goal if yes else prompt_client_approval(r2_goal, round_num=2)

    # -- Round 2 --------------------------------------------------------------
    print(f"\n{'-'*60}")
    print(f"ROUND 2 - Data capture ({len(positives)} vendor(s))")
    print(f"{'-'*60}")

    for lead in positives:
        print(f"\n[{lead.name}]")
        try:
            status_output = execute_call_pipeline(
                lead.phone, r2_goal, region=region, language=language, dry_run=dry_run,
                lat=None if skip_hours_gate else lead.lat,
                lon=None if skip_hours_gate else lead.lon,
            )
            if status_output.get("status") == "SKIPPED":
                print(f"  Skipped R2 — {status_output.get('skip_reason')}")
                continue
            call_id = status_output.get("run_id") or status_output.get("id") or "dry_run"
            extracted = extract_round2_fields(status_output)

            transcript_lines = parse_transcript_to_json(status_output)
            inference = {}
            if transcript_lines:
                print(f"  Running AI inference on {len(transcript_lines)}-line transcript...")
                try:
                    inference = infer_from_transcript(transcript_lines, product, round_num=2)
                    print(f"  Can supply: {inference.get('can_supply','?')} | "
                          f"Price: {inference.get('price_range','?')} | "
                          f"{inference.get('summary','')[:60]}")
                except Exception as ie:
                    print(f"  Inference error: {ie}")

            combined = {**extracted, **inference}
            with get_conn() as conn:
                update_lead_status(conn, lead.id, "completed", call_id, round_num=2)
                log_call(conn, lead.id, call_id, 2, status_output, combined or None)
        except Exception as e:
            print(f"  Error: {e}")

    with get_conn() as conn:
        print_results_table(conn, campaign_id)

    if export:
        export_excel(campaign_id, export)


def main():
    parser = argparse.ArgumentParser(description="Vendor Discovery & Outreach via CALL-E + OSM")
    parser.add_argument("--product", required=True, help="Product or service to source")
    parser.add_argument("--location", required=True, help="Location to search (city, region, etc.)")
    parser.add_argument("--limit", type=int, default=10, help="Max vendors to discover (default 10)")
    parser.add_argument("--region", default=None, help="Region hint for CALL-E (e.g. US, IN, GB)")
    parser.add_argument("--language", default=None, help="Language hint for CALL-E (e.g. English)")
    parser.add_argument("--dry-run", action="store_true", help="Plan calls but do not execute them")
    parser.add_argument("--yes", "-y", action="store_true", help="Skip all confirmation prompts")
    parser.add_argument("--export-excel", metavar="FILE.xlsx", default=None,
                        help="Export results to Excel after campaign completes")
    parser.add_argument("--skip-hours-gate", action="store_true",
                        help="Bypass business-hours check (for testing only)")
    args = parser.parse_args()

    run_campaign(
        product=args.product,
        location=args.location,
        limit=args.limit,
        region=args.region,
        language=args.language,
        dry_run=args.dry_run,
        yes=args.yes,
        export=args.export_excel,
        skip_hours_gate=args.skip_hours_gate,
    )


if __name__ == "__main__":
    main()
